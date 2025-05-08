from Bio import SeqIO
import pandas as pd
import os
import openpyxl

four_codons = ['GGT', 'GGC', 'GGA', 'GGG', 'CCT', 'CCC', 'CCA', 'CCG', 'GCT', 'GCC', 'GCA', 'GCG',
               'ACT', 'ACC', 'ACA', 'ACG', 'GTT', 'GTC', 'GTA', 'GTG', 'TCT', 'TCC', 'TCA', 'TCG',
               'CTT', 'CTC', 'CTA', 'CTG', 'CGT', 'CGC', 'CGA', 'CGG']

def calculate_third_base_proportions(sequence):
    counts = {'A': 0, 'T': 0, 'C': 0, 'G': 0}

    for i in range(0, len(sequence) - 2, 3):
        codon = sequence[i:i+3]
        if codon in four_codons:
            third_base = codon[2]
            counts[third_base] += 1
    
    total_a_t = counts['A'] + counts['T']
    total_g_c = counts['G'] + counts['C']
    
    proportion_a3_t3 = counts['A'] / total_a_t if total_a_t > 0 else 0
    proportion_g3_c3 = counts['G'] / total_g_c if total_g_c > 0 else 0
    
    return proportion_a3_t3, proportion_g3_c3

def process_fasta_file(file_path):
    records = SeqIO.parse(file_path, "fasta")
    data = []
    for record in records:
        seq = str(record.seq).upper()
        proportion_a3_t3, proportion_g3_c3 = calculate_third_base_proportions(seq)
        data.append({
            'Taxon': record.id,
            'A3/(A3+T3)': proportion_a3_t3,
            'G3/(G3+C3)': proportion_g3_c3
        })
    return pd.DataFrame(data)

def main():
    current_dir = os.getcwd()
    input_dir = current_dir  
    output_xlsx = os.path.join(current_dir, "PR2_index.xlsx")  
    
    writer = pd.ExcelWriter(output_xlsx, engine='openpyxl')
    for file_name in os.listdir(input_dir):
        if file_name.endswith(".fasta") or file_name.endswith(".fa"):
            gene_name = os.path.splitext(file_name)[0]
            file_path = os.path.join(input_dir, file_name)
            df = process_fasta_file(file_path)
            df.to_excel(writer, sheet_name=gene_name, index=False)
            
            wb = writer.book
            ws = wb[gene_name]
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
                    cell.border = None
                    cell.font = openpyxl.styles.Font(name='Times New Roman')

            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 14

    writer._save()

if __name__ == "__main__":
    main()
